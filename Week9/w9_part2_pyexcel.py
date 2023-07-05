import pandas as pd 
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from w9_helper import num_games

#   Load data into a dataframe
pd.set_option('display.max_columns', 500)

FILE_NAME = "videogame/vgsales.csv"
raw_df = pd.read_csv(FILE_NAME)
##############################################################
# What country had the highest sales for N64?
#   Filter df for N64. 
#   GROUP BY on country, sort descending.
#   Transposed the Groupby Object
#   Renamed columns
##############################################################

n64_df = raw_df[raw_df['Platform'] == 'N64']
n64_df = n64_df.groupby(['Platform'])[['NA_Sales', 'EU_Sales', 'JP_Sales']].sum().reset_index().T.reset_index()

#   .iloc --- "SELECTION" using integer
#   .loc  --- "SELECTION" using label
n64_df = n64_df.iloc[1:, :] 

n64_df.columns = ['Platform', 'Sales']
n64_df = n64_df.sort_values(['Sales'], ascending=False)

wb = Workbook() #   Created workbook
del wb['Sheet']
ws1 = wb.create_sheet('Question1')

for r in dataframe_to_rows(n64_df, header=True, index=False): 
    ws1.append(r)

##############################################################
#   Question 2
# - #### What is the distribution of the genres for the top 100 games?
#   - How many unique genres are there and how would you visualize this?
##############################################################
ws2 = wb.create_sheet('Question2')

highest_sales_df = raw_df.sort_values(['Global_Sales'], ascending=False)
highest_sales_df = highest_sales_df.iloc[:100, :]

genre_counts_df = highest_sales_df.value_counts(['Genre', 'Platform']).reset_index()
genre_counts_df = genre_counts_df.sort_values(['Genre'])

for r in dataframe_to_rows(genre_counts_df, header=True, index=False): 
    ws2.append(r)

##############################################################
#   Question 3
# - #### Create a `Decade` column and visualise this with the appropriate visual.
#   - What visual did you choose and why? 
##############################################################

ws3 = wb.create_sheet('Question3')

# df.loc[(condition), new_col_name] = value
raw_df.loc[(raw_df['Year'] < 1990), 'Decade'] = "pre_1990"
raw_df.loc[((raw_df['Year'] >= 1990) & (raw_df['Year'] < 2000)), 'Decade'] = "1990s"
raw_df.loc[((raw_df['Year'] >= 2000) & (raw_df['Year'] < 2010)), 'Decade'] = "2000s"
raw_df.loc[((raw_df['Year'] >= 2010) & (raw_df['Year'] < 2020)), 'Decade'] = "2010s"

decade_df = num_games(raw_df, "Decade", "Name")

for r in dataframe_to_rows(decade_df, header=True, index=False): 
    ws3.append(r)

wb.save('Week9/Week9P2.xlsx')






# ### Create a file called `week9_part2.py` to answer the questions above. 

# ### - BONUS: Create a function to do this and put it in a separate file. 